import io
import re
import math
import zipfile
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional, Any

import pandas as pd
import streamlit as st

from mstower_geometry import (
    FAMILY_ORDER as GEOMETRY_FAMILY_ORDER,
    build_tower_geometry,
    geometry_members_dataframe,
    geometry_panels_dataframe,
    material_schedule_dataframe,
    geometry_validation_dataframe,
    geometry_validation_summary,
    make_engineering_2d_figure,
    make_material_elevation_figure,
    make_tower_3d_figure,
)

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = None


# =========================================================
# APP CONFIG
# =========================================================
st.set_page_config(
    page_title="MS Tower Batch Extractor Pro",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

APP_TITLE = "🏗️ MS Tower Batch Extractor Pro"
APP_SUBTITLE = (
    "Batch parser untuk output MS Tower dengan fokus pada keputusan engineering: "
    "governing ratio, displacement, sway/twist, support reaction, dan export Excel."
)

SERVICE_KEYWORDS = ("OPERATIONAL",)
DISP_SECTION_START = (
    "NODE DISPLACEMENTS",
    "N O D E   D I S P L A C E M E N T S",
)
DISP_SECTION_END = (
    "ENVELOPE",
    "M E M B E R",
    "SUPPORT",
    "TOWER ROTATIONS",
    "REACTIONS",
)
ROTATION_START = "ENVELOPE OF TOWER ROTATIONS"


# =========================================================
# DATA MODELS
# =========================================================
@dataclass
class ParseResult:
    filename: str
    tower_type: str = "Unknown"
    total_height_m: float = 0.0
    dfs: Dict[str, pd.DataFrame] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)

    @property
    def success(self) -> bool:
        return bool(self.dfs) and not self.errors

    @property
    def status(self) -> str:
        if self.errors:
            return "❌ Error"
        if self.warnings:
            return "⚠️ Success / Review"
        return "✅ Success"


# =========================================================
# HELPERS
# =========================================================
def decode_rpt(raw: bytes) -> str:
    """Decode common MS Tower text encodings without failing the batch."""
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def split_klr(val: Any) -> Tuple[Any, str]:
    match = re.match(r"^([0-9.]+)([a-zA-Z]+)?$", str(val).strip())
    if match:
        return match.group(1), match.group(2) or ""
    return val, ""


def num(value: Any) -> Optional[float]:
    """Safe numeric parser for report tokens."""
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def safe_round(value: Any, ndigits: int = 4):
    if value is None or pd.isna(value):
        return None
    return round(float(value), ndigits)


def status_label(value: Any, limit: Any) -> str:
    v = num(value)
    lim = num(limit)
    if v is None or lim is None or lim <= 0:
        return "N/A"
    return "PASS" if v <= lim else "FAIL"


def max_abs_numeric(df: pd.DataFrame, col: str):
    if col not in df.columns:
        return None, None
    s = pd.to_numeric(df[col], errors="coerce")
    if s.notna().sum() == 0:
        return None, None
    idx = s.abs().idxmax()
    return float(s.loc[idx]), idx


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    return out


def df_to_excel_bytes(extracted_dfs: Dict[str, pd.DataFrame]) -> bytes:
    """Create a formatted workbook only when requested/needed."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        preferred = [
            "Overall Check",
            "Resume",
            "Pole Resume",
            "Load Parameters",
            "Antenna Height Resume",
            "Antenna Loading",
            "Linear Loading",
            "Tower Profile",
            "Tower Member Resume",
            "Tower Member Detail",
            "Member Section Catalog",
            "Bolt Data",
            "Reaction Resume",
            "Data SR",
            "Pole SR",
            "Mass Summary",
            "Support Reactions",
        ]

        ordered = [x for x in preferred if x in extracted_dfs]
        ordered += [x for x in extracted_dfs.keys() if x not in ordered]

        for sheet in ordered:
            df = normalize_columns(extracted_dfs[sheet])
            safe_sheet = str(sheet)[:31]
            df.to_excel(writer, index=False, sheet_name=safe_sheet)

        wb = writer.book
        thin = Side(style="thin", color="D9E1F2") if Side else None

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            if Font and PatternFill:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1F4E78")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            if thin:
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = Border(bottom=thin)
                        cell.alignment = Alignment(vertical="center")

            if get_column_letter:
                for col_idx in range(1, ws.max_column + 1):
                    values = [
                        len(str(ws.cell(row=r, column=col_idx).value or ""))
                        for r in range(1, min(ws.max_row, 100) + 1)
                    ]
                    width = min(max(max(values, default=10) + 2, 10), 36)
                    ws.column_dimensions[get_column_letter(col_idx)].width = width

    return output.getvalue()


# =========================================================
# LOAD / PARAMETER PARSERS
# =========================================================
def _clean_source_line(line: str) -> str:
    """Remove non-printing control bytes that can appear in MST/TWR text blocks."""
    return "".join(ch for ch in str(line) if ch in ("\t",) or ord(ch) >= 32).strip()


def _source_lines(content: str) -> List[str]:
    """
    Normalize text/binary-ish MST content into reliable logical lines.

    Some .mst/.twr exports contain CR-only line endings and control-byte
    separators (for example \x04). Converting those separators to newlines
    makes PARAMETERS / ANCILLARIES / LINEAR blocks parse consistently.
    """
    text = str(content).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]+", "\n", text)
    return [_clean_source_line(line) for line in text.split("\n")]


def _token_after(tokens: List[str], key: str) -> Optional[str]:
    """Return the token immediately after a keyword, case-insensitively."""
    key_u = key.upper()
    for idx, token in enumerate(tokens[:-1]):
        if str(token).upper() == key_u:
            return tokens[idx + 1]
    return None


def _to_number_if_possible(value: Optional[str]):
    if value is None:
        return None
    parsed = num(value)
    return parsed if parsed is not None else str(value).strip()


def _status_from_group(group: str) -> str:
    upper = str(group).upper()
    if "EXISTING" in upper:
        return "Existing"
    if "PROPOSED" in upper:
        return "Proposed"
    return ""


def _tenant_from_group(group: str) -> str:
    """Use text after ANTENNA as tenant/owner when the row has no inline remark."""
    match = re.search(r"\bANTENNA\b\s+(.+)$", str(group), re.I)
    return match.group(1).strip() if match else ""


def _equipment_type(item: str, group: str = "", library: str = "") -> str:
    """Classify tower-mounted equipment using item name, group and LIB token.

    MS Tower projects are not always consistent in naming: the same physical
    device may be identifiable from the item prefix (AAU1, ODU2, FILTER3) or
    only from its library code.  The classifier therefore evaluates all three
    text sources and keeps the more specific device types ahead of generic RF/MW.
    """
    upper_item = str(item).upper().strip()
    upper_group = str(group).upper().strip()
    upper_library = str(library).upper().strip()
    combined = f"{upper_item} {upper_library} {upper_group}"

    # LINEAR objects
    if upper_item.startswith("LADDER") or "LADDER" in upper_group:
        return "Ladder"
    if upper_item.startswith("TRAY") or "TRAY" in upper_group:
        return "Tray"
    if upper_item.startswith("FEEDER") or "FEEDER" in upper_group:
        return "Feeder"

    # Active / antenna-mounted equipment. More specific types are tested first.
    if upper_item.startswith("AAU") or re.search(r"\bAAU", combined):
        return "AAU"
    if upper_item.startswith("RRU") or re.search(r"\bRRU", combined):
        return "RRU"
    if upper_item.startswith("RRH") or re.search(r"\bRRH", combined):
        return "RRH"
    if upper_item.startswith("ODU") or "ODU" in upper_item or re.search(r"\bODU", combined):
        return "ODU"
    if (
        upper_item.startswith(("FILTER", "FLT"))
        or "FILTER" in combined
        or re.search(r"\bFLT", combined)
    ):
        return "Filter"
    if upper_item.startswith("TMA") or re.search(r"\bTMA", combined):
        return "TMA"
    if upper_item.startswith(("COMB", "COMBINER")) or "COMBINER" in combined:
        return "Combiner"
    if upper_item.startswith(("DIP", "DIPLEX")) or "DIPLEX" in combined:
        return "Diplexer"
    if upper_item.startswith(("TRIP", "TRIPLEX")) or "TRIPLEX" in combined:
        return "Triplexer"
    if upper_item.startswith("GPS") or re.search(r"\bGPS", combined):
        return "GPS"

    # Microwave antenna is intentionally below ODU so e.g. MWODU / ODU library
    # is not incorrectly counted as a dish antenna.
    if upper_item.startswith("MW") or re.search(r"\b(MW|MICROWAVE)", combined):
        return "Microwave"

    # Sector/panel/RF antenna naming can be carried by item or library code.
    if (
        upper_item.startswith(("RF", "ANT", "SECTOR", "PANEL"))
        or upper_library.startswith(("SEC", "RF", "ANT"))
        or "SECTOR ANT" in combined
    ):
        return "RF Antenna"

    return "Other"


def extract_load_parameters(content: str) -> pd.DataFrame:
    """Extract the PARAMETERS block into a structured key/value table."""
    rows: List[Dict[str, Any]] = []
    capturing = False

    for raw_line in _source_lines(content):
        line = _clean_source_line(raw_line)
        if not line:
            continue

        upper = line.upper()
        if not capturing:
            pos = upper.find("PARAMETERS")
            if pos < 0:
                continue
            capturing = True
            line = line[pos + len("PARAMETERS"):].strip()
            if not line:
                continue
            upper = line.upper()

        if re.match(r"^END\b", upper):
            break
        if line.startswith("$"):
            continue

        data_part, _, comment = line.partition("$")
        parts = data_part.split()
        if len(parts) < 2:
            continue

        parameter = parts[0].strip()
        value_text = " ".join(parts[1:]).strip()
        value_num = num(value_text)
        value = value_num if value_num is not None else value_text

        rows.append(
            {
                "Parameter": parameter,
                "Value": value,
                "Description": comment.strip(),
            }
        )

    return pd.DataFrame(rows, columns=["Parameter", "Value", "Description"])


def extract_antenna_loading(content: str) -> pd.DataFrame:
    """Extract ANCILLARIES antenna/device rows such as MW, RF, RRU and GPS."""
    rows: List[Dict[str, Any]] = []
    capturing = False
    group = ""

    for raw_line in _source_lines(content):
        line = _clean_source_line(raw_line)
        if not line:
            continue

        upper = line.upper()
        if not capturing:
            if re.search(r"\bANCILLARIES\b", upper):
                capturing = True
            continue

        if re.match(r"^FACE\b", upper):
            break

        if line.startswith("$"):
            group = line[1:].strip()
            continue

        if upper.startswith("LARGE LIBR") or upper.startswith("LINEAR LIBR"):
            continue
        if re.match(r"^END\b", upper):
            break

        data_part, _, inline_comment = line.partition("$")
        tokens = data_part.split()
        if not tokens:
            continue

        required = ("XA", "YA", "ZA", "LIB", "ANG")
        if not all(any(str(tok).upper() == key for tok in tokens) for key in required):
            continue

        item = tokens[0]
        library = _token_after(tokens, "LIB") or ""
        rows.append(
            {
                "Status": _status_from_group(group),
                "Load Group": group,
                "Item": item,
                "Equipment Type": _equipment_type(item, group, library),
                "XA (m)": _to_number_if_possible(_token_after(tokens, "XA")),
                "YA (m)": _to_number_if_possible(_token_after(tokens, "YA")),
                "ZA (m)": _to_number_if_possible(_token_after(tokens, "ZA")),
                "Library": library,
                "Angle (deg)": _to_number_if_possible(_token_after(tokens, "ANG")),
                "Added Mass": _to_number_if_possible(_token_after(tokens, "AMASS")),
                "Tenant / Remark": inline_comment.strip() or _tenant_from_group(group),
            }
        )

    columns = [
        "Status", "Load Group", "Item", "Equipment Type",
        "XA (m)", "YA (m)", "ZA (m)", "Library",
        "Angle (deg)", "Added Mass", "Tenant / Remark",
    ]
    return pd.DataFrame(rows, columns=columns)


def extract_linear_loading(content: str) -> pd.DataFrame:
    """Extract LINEAR loads such as ladder, tray and feeder robustly."""
    rows: List[Dict[str, Any]] = []
    capturing = False
    group = ""
    current_status = ""

    for raw_line in _source_lines(content):
        line = _clean_source_line(raw_line)
        if not line:
            continue

        upper = line.upper()

        # LINEAR LIBR is the reliable start marker. FACE may be on its own
        # line, on the same logical line, or separated by binary control bytes.
        if not capturing:
            if re.search(r"\bLINEAR\s+LIBR\b", upper):
                capturing = True
            continue

        if re.match(r"^END\b", upper):
            break

        if line.startswith("$"):
            group = line[1:].strip()
            parsed_status = _status_from_group(group)
            if parsed_status:
                current_status = parsed_status
            continue

        if "LINEAR LIBR" in upper or "LARGE LIBR" in upper:
            continue

        data_part, _, inline_comment = line.partition("$")
        tokens = data_part.split()
        if not tokens:
            continue

        # Linear objects all have bottom/top XYZ plus a library definition.
        # FACT is optional (e.g. LADDER often has no FACT token).
        required = ("XB", "YB", "ZB", "XT", "YT", "ZT", "LIB")
        token_keys = {str(tok).upper() for tok in tokens}
        if not all(key in token_keys for key in required):
            continue

        item = tokens[0]
        load_type = _equipment_type(item, group)

        # Avoid accidentally accepting another LINEAR object type that is not
        # part of the requested load inventory.
        if load_type not in {"Ladder", "Tray", "Feeder"}:
            continue

        rows.append(
            {
                "Status": current_status or _status_from_group(group),
                "Load Group": group,
                "Item": item,
                "Load Type": load_type,
                "XB (m)": _to_number_if_possible(_token_after(tokens, "XB")),
                "YB (m)": _to_number_if_possible(_token_after(tokens, "YB")),
                "ZB (m)": _to_number_if_possible(_token_after(tokens, "ZB")),
                "XT (m)": _to_number_if_possible(_token_after(tokens, "XT")),
                "YT (m)": _to_number_if_possible(_token_after(tokens, "YT")),
                "ZT (m)": _to_number_if_possible(_token_after(tokens, "ZT")),
                "Library": _token_after(tokens, "LIB") or "",
                "Factor": _to_number_if_possible(_token_after(tokens, "FACT")),
                "Angle (deg)": _to_number_if_possible(_token_after(tokens, "ANG")),
                "Remark": inline_comment.strip(),
            }
        )

    columns = [
        "Status", "Load Group", "Item", "Load Type",
        "XB (m)", "YB (m)", "ZB (m)",
        "XT (m)", "YT (m)", "ZT (m)",
        "Library", "Factor", "Angle (deg)", "Remark",
    ]
    return pd.DataFrame(rows, columns=columns)


def _device_summary_label(value: Any) -> str:
    """Compact *original* device label used in textual summaries.

    The resume table intentionally groups FILTER/TMA/COMBINER/DIPLEXER/
    TRIPLEXER/GPS (and unforeseen equipment) into one numeric ``Others``
    column.  Text summaries, however, retain the original classified device
    name so the engineering meaning is not lost.
    """
    text = str(value).strip().upper()
    mapping = {
        "RF ANTENNA": "RF",
        "MICROWAVE": "MW",
        "FILTER": "FILTER",
        "COMBINER": "COMBINER",
        "DIPLEXER": "DIPLEXER",
        "TRIPLEXER": "TRIPLEXER",
        "OTHER": "OTHER",
    }
    return mapping.get(text, text or "OTHER")


def _device_resume_bucket(value: Any) -> str:
    """Map equipment to the numeric columns shown in the elevation resume."""
    label = _device_summary_label(value)
    dedicated = {"RF", "RRU", "RRH", "AAU", "MW", "ODU"}
    return label if label in dedicated else "Others"


# Only these receive their own numeric columns.  All other equipment is
# aggregated into ``Others`` while preserving its actual name in summaries.
DEVICE_RESUME_ORDER = ["RF", "RRU", "RRH", "AAU", "MW", "ODU", "Others"]

# Ordering for textual summaries.  The names below remain visible even when
# their numeric count is rolled up into the Others column.
DEVICE_TEXT_ORDER = [
    "RF", "RRU", "RRH", "AAU", "MW", "ODU",
    "FILTER", "TMA", "COMBINER", "DIPLEXER", "TRIPLEXER", "GPS", "OTHER",
]


def _ordered_device_text_types(values) -> List[str]:
    """Stable engineering-friendly order for original device names."""
    present = {str(v).strip() for v in values if str(v).strip()}
    ordered = [x for x in DEVICE_TEXT_ORDER if x in present]
    ordered += sorted(x for x in present if x not in DEVICE_TEXT_ORDER)
    return ordered


def _device_count_text(df: pd.DataFrame) -> str:
    """Return compact counts using original device names.

    Example: ``3 RF + 6 RRU + 2 FILTER + 1 GPS``.  FILTER and GPS still roll
    up into the numeric ``Others`` column, but their original names remain
    visible here and in the per-tenant summary.
    """
    if df.empty:
        return ""

    if "Device Label" in df.columns:
        types = df["Device Label"].astype(str)
    else:
        types = df["Equipment Type"].apply(_device_summary_label).astype(str)

    counts = types.value_counts()
    ordered = _ordered_device_text_types(counts.index)
    return " + ".join(f"{int(counts[t])} {t}" for t in ordered)


def build_antenna_height_resume(antenna_loading: pd.DataFrame) -> pd.DataFrame:
    """Build device resume per elevation with compact engineering columns.

    Dedicated numeric columns are provided for RF, RRU, RRH, AAU, MW and ODU.
    FILTER, TMA, COMBINER, DIPLEXER, TRIPLEXER, GPS and any unforeseen device
    type are accumulated in a single ``Others`` column.  ``Device Summary``
    and ``Tenant Device Summary`` retain the original equipment names.

    ZA is grouped into its integer-metre elevation band (floor for positive
    tower elevations). Thus ZA 60.20 and ZA 60.50 are summarized at 60 m.
    """
    base_columns = ["Elevation (m)", "Total Device"]
    tail_columns = ["Device Summary", "Tenant Device Summary", "Azimuth", "Status Summary"]

    if antenna_loading.empty or "ZA (m)" not in antenna_loading.columns:
        return pd.DataFrame(columns=base_columns + DEVICE_RESUME_ORDER + tail_columns)

    work = antenna_loading.copy()
    work["ZA Numeric"] = pd.to_numeric(work["ZA (m)"], errors="coerce")
    work = work.dropna(subset=["ZA Numeric"])
    if work.empty:
        return pd.DataFrame(columns=base_columns + DEVICE_RESUME_ORDER + tail_columns)

    work["Elevation (m)"] = work["ZA Numeric"].apply(lambda x: math.floor(float(x)))
    work["Device Label"] = work["Equipment Type"].apply(_device_summary_label)
    work["Resume Bucket"] = work["Equipment Type"].apply(_device_resume_bucket)
    work["Tenant Clean"] = (
        work.get("Tenant / Remark", pd.Series(index=work.index, dtype=object))
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", "Unspecified")
    )

    # Keep the resume compact: only show numeric device columns that are
    # actually present, with Others as one aggregate column when required.
    present_buckets = set(work["Resume Bucket"].astype(str))
    device_columns = [x for x in DEVICE_RESUME_ORDER if x in present_buckets]
    columns = base_columns + device_columns + tail_columns

    out_rows: List[Dict[str, Any]] = []
    for elevation, elev_df in work.groupby("Elevation (m)", sort=True):
        bucket_counts = elev_df["Resume Bucket"].value_counts()

        tenant_parts = []
        for tenant, tenant_df in elev_df.groupby("Tenant Clean", sort=True):
            device_text = _device_count_text(tenant_df)
            tenant_parts.append(f"{device_text} ({tenant})" if device_text else str(tenant))

        azimuth_text = "—"
        if "Angle (deg)" in elev_df.columns:
            angle_num = pd.to_numeric(elev_df["Angle (deg)"], errors="coerce").dropna()
            if not angle_num.empty:
                unique_angles = sorted({float(v) % 360.0 for v in angle_num.tolist()})
                def _fmt_azimuth(v: float) -> str:
                    return f"{int(v)}°" if float(v).is_integer() else f"{v:g}°"
                azimuth_text = ", ".join(_fmt_azimuth(v) for v in unique_angles)

        status_parts = []
        if "Status" in elev_df.columns:
            status_counts = elev_df["Status"].fillna("").astype(str).str.strip().value_counts()
            for status, count in status_counts.items():
                if status:
                    status_parts.append(f"{int(count)} {status}")

        row: Dict[str, Any] = {
            "Elevation (m)": int(elevation),
            "Total Device": len(elev_df),
            "Device Summary": _device_count_text(elev_df),
            "Tenant Device Summary": " | ".join(tenant_parts),
            "Azimuth": azimuth_text,
            "Status Summary": " | ".join(status_parts),
        }
        for device_type in device_columns:
            row[device_type] = int(bucket_counts.get(device_type, 0))
        out_rows.append(row)

    return pd.DataFrame(out_rows, columns=columns).sort_values(
        "Elevation (m)", ascending=False, ignore_index=True
    )


# =========================================================
# TOWER MEMBER / SECTION / BOLT PARSERS
# =========================================================
def extract_member_section_catalog(content: str) -> pd.DataFrame:
    """Extract the SECTIONS block used as the member-size master lookup.

    Example::
        70  EA70x70x7    fy 245
        66  DAE60X60X6   fy 245

    The numeric section ID is what PROFILE lines reference in expressions such
    as ``LEG 70`` or ``BR1 50``.
    """
    rows: List[Dict[str, Any]] = []
    capturing = False
    category = "Standard"
    library = ""
    ifact = None

    for raw_line in _source_lines(content):
        line = _clean_source_line(raw_line)
        if not line:
            continue
        upper = line.upper()

        if not capturing:
            if re.match(r"^SECTIONS\b", upper):
                capturing = True
            continue

        if re.match(r"^END\b", upper):
            break

        if line.startswith("$"):
            comment = line[1:].strip()
            if "STRENGTHENED" in comment.upper():
                category = "Strengthened"
            continue

        if upper.startswith("LIBR "):
            tokens = line.split()
            if len(tokens) >= 2:
                library = tokens[1]
            ifact = _to_number_if_possible(_token_after(tokens, "IFACT"))
            continue

        match = re.match(r"^(\d+)\s+(\S+)(.*)$", line)
        if not match:
            continue

        section_id = int(match.group(1))
        section_name = match.group(2).strip()
        remainder = match.group(3)
        fy_match = re.search(r"\bfy\s+([-+]?\d+(?:\.\d+)?)", remainder, re.I)
        fy = float(fy_match.group(1)) if fy_match else None

        rows.append(
            {
                "Section ID": section_id,
                "Section / Profile": section_name,
                "fy": fy,
                "Category": category,
                "Library": library,
                "IFACT": ifact,
            }
        )

    return pd.DataFrame(
        rows,
        columns=["Section ID", "Section / Profile", "fy", "Category", "Library", "IFACT"],
    )


def _member_family(line_type: str, member_key: str) -> str:
    """Convert PROFILE member tokens into compact engineering families."""
    source = str(line_type).upper().strip()
    key = str(member_key).upper().strip()

    if key == "LEG":
        return "LEG"
    if source == "PLAN" and re.match(r"^PB\d+$", key):
        return "PBR"
    if source == "HIP" and re.match(r"^HP\d+$", key):
        return "HIP"
    if source == "FACE":
        if re.match(r"^BR\d+$", key):
            return "BR"
        if re.match(r"^H\d+$", key):
            return "HOR"
        if re.match(r"^R\d+$", key):
            return "RED"
    return key


def _is_profile_member_key(line_type: str, token: str) -> bool:
    source = str(line_type).upper().strip()
    key = str(token).upper().strip()
    if source == "FACE":
        return bool(re.match(r"^(?:LEG|BR\d+|H\d+|R\d+)$", key))
    if source == "PLAN":
        return bool(re.match(r"^PB\d+$", key))
    if source == "HIP":
        return bool(re.match(r"^HP\d+$", key))
    return False


def _format_section_assignment(group: pd.DataFrame) -> str:
    """Compact repeated member keys that share the same section ID/profile."""
    if group.empty:
        return ""

    parts: List[str] = []
    # Preserve first-seen order instead of alphabetically sorting keys such as R10 before R2.
    seen_section_keys: List[Tuple[Any, str, Any]] = []
    for _, row in group.iterrows():
        signature = (
            row.get("Section ID"),
            str(row.get("Section / Profile", "")),
            row.get("fy"),
        )
        if signature not in seen_section_keys:
            seen_section_keys.append(signature)

    for section_id, profile, fy in seen_section_keys:
        same = group[
            (group["Section ID"] == section_id)
            & (group["Section / Profile"].astype(str) == str(profile))
        ]
        key_values = []
        for key_value in same["Member Key"].astype(str).tolist():
            if key_value not in key_values:
                key_values.append(key_value)
        keys = "/".join(key_values)
        profile_text = profile if profile and profile != "nan" else "Unmapped"
        fy_text = "" if fy is None or pd.isna(fy) else f", fy {fy:g}"
        parts.append(f"{keys} = {int(section_id)} {profile_text}{fy_text}")
    return " ; ".join(parts)


def extract_tower_member_data(
    content: str, section_catalog: pd.DataFrame
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Parse PROFILE and resolve every non-zero member code through SECTIONS.

    Returns ``(member_detail, member_resume, tower_profile)``.
    """
    section_lookup: Dict[str, Dict[str, Any]] = {}
    if not section_catalog.empty:
        for _, row in section_catalog.iterrows():
            sid = row.get("Section ID")
            if pd.notna(sid):
                section_lookup[str(int(float(sid)))] = row.to_dict()

    profile_meta: Dict[str, Any] = {
        "Title 1": "",
        "Title 2": "",
        "Units": "",
        "Faces": None,
        "WBASE": None,
        "RLBAS": None,
    }

    rows: List[Dict[str, Any]] = []
    panels: Dict[int, Dict[str, Any]] = {}
    capturing = False
    current_low = None
    current_high = None
    current_panel: Optional[int] = None

    logical_lines = _source_lines(content)

    # General project/profile metadata.
    for line in logical_lines:
        clean = _clean_source_line(line)
        upper = clean.upper()
        if upper.startswith("TITL1 "):
            profile_meta["Title 1"] = clean[6:].strip()
        elif upper.startswith("TITL2 "):
            profile_meta["Title 2"] = clean[6:].strip()
        elif re.match(r"^UNITS\s+", upper):
            profile_meta["Units"] = clean.split()[1] if len(clean.split()) > 1 else ""
        elif re.match(r"^FACES\s+", upper):
            profile_meta["Faces"] = _to_number_if_possible(clean.split()[1])
        elif re.match(r"^WBASE\s+", upper):
            profile_meta["WBASE"] = _to_number_if_possible(clean.split()[1])
        elif re.match(r"^RLBAS\s+", upper):
            profile_meta["RLBAS"] = _to_number_if_possible(clean.split()[1])

    for raw_line in logical_lines:
        line = _clean_source_line(raw_line)
        if not line:
            continue
        upper = line.upper()

        if not capturing:
            if re.match(r"^PROFILE\b", upper):
                capturing = True
            continue

        if re.match(r"^END\b", upper):
            break

        if line.startswith("$"):
            section_match = re.search(
                r"\bSECTION\s+([-+]?\d+(?:\.\d+)?)\s*-\s*([-+]?\d+(?:\.\d+)?)\s*M\b",
                line,
                re.I,
            )
            if section_match:
                current_low = float(section_match.group(1))
                current_high = float(section_match.group(2))
            continue

        panel_match = re.match(
            r"^PANEL\s+(\d+)\s+HT\s+([-+]?\d+(?:\.\d+)?)(?:\s+TW\s+([-+]?\d+(?:\.\d+)?))?",
            line,
            re.I,
        )
        if panel_match:
            current_panel = int(panel_match.group(1))
            panels[current_panel] = {
                "Panel": current_panel,
                "Section From (m)": current_low,
                "Section To (m)": current_high,
                "Panel HT (m)": float(panel_match.group(2)),
                "TW": float(panel_match.group(3)) if panel_match.group(3) else None,
                "FACE Pattern": "",
                "PLAN Pattern": "",
                "HIP Pattern": "",
            }
            continue

        if current_panel is None:
            continue

        tokens = line.split()
        if len(tokens) < 2:
            continue
        line_type = tokens[0].upper()
        if line_type not in {"FACE", "PLAN", "HIP"}:
            continue

        pattern = tokens[1]
        qualifier_tokens = [tok.upper() for tok in tokens[2:] if tok.upper() in {"TOP", "XIP"}]
        qualifier = "/".join(dict.fromkeys(qualifier_tokens))
        pattern_label = f"{pattern} {qualifier}".strip()
        pattern_key = f"{line_type} Pattern"
        existing_pattern = str(panels[current_panel].get(pattern_key, "") or "")
        existing_parts = [x.strip() for x in existing_pattern.split(" | ") if x.strip()]
        if pattern_label not in existing_parts:
            existing_parts.append(pattern_label)
        panels[current_panel][pattern_key] = " | ".join(existing_parts)

        idx = 2
        while idx < len(tokens):
            key = tokens[idx].upper()
            if key in {"TOP", "XIP"}:
                idx += 1
                continue
            if idx + 1 >= len(tokens):
                break

            value_token = tokens[idx + 1]
            section_value = num(value_token)
            if _is_profile_member_key(line_type, key) and section_value is not None:
                section_id = int(section_value)
                # A zero member code explicitly means no section/member at this slot.
                if section_id > 0:
                    mapped = section_lookup.get(str(section_id), {})
                    profile = mapped.get("Section / Profile", "")
                    fy = mapped.get("fy")
                    category = mapped.get("Category", "")
                    rows.append(
                        {
                            "Panel": current_panel,
                            "Section From (m)": current_low,
                            "Section To (m)": current_high,
                            "Elevation Range": (
                                f"{current_low:g} - {current_high:g} m"
                                if current_low is not None and current_high is not None
                                else ""
                            ),
                            "Panel HT (m)": panels[current_panel].get("Panel HT (m)"),
                            "TW": panels[current_panel].get("TW"),
                            "Line Type": line_type,
                            "Pattern": pattern,
                            "Qualifier": qualifier,
                            "Member Family": _member_family(line_type, key),
                            "Member Key": key,
                            "Section ID": section_id,
                            "Section / Profile": profile,
                            "fy": fy,
                            "Section Category": category,
                            "Mapped": "Yes" if mapped else "No",
                        }
                    )
                idx += 2
            else:
                idx += 1

    detail_columns = [
        "Panel", "Section From (m)", "Section To (m)", "Elevation Range",
        "Panel HT (m)", "TW", "Line Type", "Pattern", "Qualifier", "Member Family",
        "Member Key", "Section ID", "Section / Profile", "fy",
        "Section Category", "Mapped",
    ]
    detail = pd.DataFrame(rows, columns=detail_columns)
    if not detail.empty:
        detail = detail.sort_values(
            ["Panel", "Line Type", "Member Family", "Member Key"],
            kind="stable",
            ignore_index=True,
        )

    family_order = ["LEG", "BR", "HOR", "RED", "PBR", "HIP"]
    resume_rows: List[Dict[str, Any]] = []
    for panel_no in sorted(panels):
        info = panels[panel_no]
        panel_df = detail[detail["Panel"] == panel_no] if not detail.empty else pd.DataFrame()
        row: Dict[str, Any] = {
            "Panel": panel_no,
            "Elevation Range": (
                f"{info['Section From (m)']:g} - {info['Section To (m)']:g} m"
                if info.get("Section From (m)") is not None and info.get("Section To (m)") is not None
                else ""
            ),
            "HT (m)": info.get("Panel HT (m)"),
            "TW": info.get("TW"),
            "FACE Pattern": info.get("FACE Pattern", ""),
            "PLAN Pattern": info.get("PLAN Pattern", ""),
            "HIP Pattern": info.get("HIP Pattern", ""),
        }
        for family in family_order:
            fam_df = panel_df[panel_df["Member Family"] == family] if not panel_df.empty else pd.DataFrame()
            row[family] = _format_section_assignment(fam_df) if not fam_df.empty else "—"
        resume_rows.append(row)

    resume_columns = [
        "Panel", "Elevation Range", "HT (m)", "TW",
        "FACE Pattern", "PLAN Pattern", "HIP Pattern",
    ] + family_order
    resume = pd.DataFrame(resume_rows, columns=resume_columns)

    profile_meta["Total Panels"] = len(panels)
    profile_meta["Member Assignments"] = len(detail)
    profile_meta["Section Types"] = len(section_catalog)
    tower_profile = pd.DataFrame([profile_meta])

    return detail, resume, tower_profile


def extract_bolt_data(content: str) -> pd.DataFrame:
    """Extract BOLTDATA definitions into a structured table."""
    rows: List[Dict[str, Any]] = []
    capturing = False

    for raw_line in _source_lines(content):
        line = _clean_source_line(raw_line)
        if not line:
            continue
        upper = line.upper()

        if not capturing:
            if re.match(r"^BOLTDATA\b", upper):
                capturing = True
            continue

        if re.match(r"^END\b", upper):
            break
        if line.startswith("$"):
            continue

        data_part, _, comment = line.partition("$")
        tokens = data_part.split()
        if len(tokens) < 2:
            continue

        bolt_id = tokens[0]
        grade = tokens[1]
        # Bolt rows are expected to carry a D token; this prevents unrelated
        # text from being accepted if a variant report embeds notes in BOLTDATA.
        if _token_after(tokens, "D") is None:
            continue

        rows.append(
            {
                "Bolt ID": bolt_id,
                "Grade": grade,
                "Diameter (mm)": _to_number_if_possible(_token_after(tokens, "D")),
                "AS": _to_number_if_possible(_token_after(tokens, "AS")),
                "FY": _to_number_if_possible(_token_after(tokens, "FY")),
                "FU": _to_number_if_possible(_token_after(tokens, "FU")),
                "FV": _to_number_if_possible(_token_after(tokens, "FV")),
                "X": _to_number_if_possible(_token_after(tokens, "X")),
                "Y": _to_number_if_possible(_token_after(tokens, "Y")),
                "Z": _to_number_if_possible(_token_after(tokens, "Z")),
                "NSP": _to_number_if_possible(_token_after(tokens, "NSP")),
                "LJ": _to_number_if_possible(_token_after(tokens, "LJ")),
                "Remark": comment.strip(),
            }
        )

    columns = [
        "Bolt ID", "Grade", "Diameter (mm)", "AS", "FY", "FU", "FV",
        "X", "Y", "Z", "NSP", "LJ", "Remark",
    ]
    return pd.DataFrame(rows, columns=columns)


# =========================================================
# PARSER
# =========================================================
@st.cache_data(show_spinner=False)
def parse_rpt_cached(content: str) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Any]]:
    lines = content.splitlines()
    upper_lines = [line.upper() for line in lines]

    dfs: Dict[str, pd.DataFrame] = {}
    warnings: List[str] = []

    # -----------------------------------------------------
    # 0. PANEL HEIGHT / ELEVATION
    # -----------------------------------------------------
    panel_ht: Dict[int, float] = {}
    panel_pattern = re.compile(r"PANEL\s+(\d+)\s+HT\s+([0-9.]+)", re.I)

    for line in lines:
        m = panel_pattern.search(line)
        if m:
            panel_ht[int(m.group(1))] = float(m.group(2))

    cumulative_elev: Dict[str, float] = {}
    if panel_ht:
        current = 0.0
        for pnl in sorted(panel_ht.keys(), reverse=True):
            current += panel_ht[pnl]
            cumulative_elev[str(pnl)] = round(current, 2)

    # -----------------------------------------------------
    # 1. SERVICE CASES
    # -----------------------------------------------------
    service_cases = set()
    current_combin_tracking = None
    case_pattern = re.compile(r"(?:COMBIN|CASE)\s+(\d+)(.*)", re.I)

    for line in lines:
        m = case_pattern.search(line)
        if m:
            current_combin_tracking = m.group(1)
            desc = m.group(2).upper()
            if any(k in desc for k in SERVICE_KEYWORDS):
                service_cases.add(current_combin_tracking)
        if current_combin_tracking and re.search(r"\b0\.49\b", line):
            service_cases.add(current_combin_tracking)

    # -----------------------------------------------------
    # 2. NODE DISPLACEMENTS
    # -----------------------------------------------------
    max_disp = 0.0
    case_disp = ""
    in_disp_section = False
    current_case_header = None

    for line, upper in zip(lines, upper_lines):
        if (
            any(marker in upper for marker in DISP_SECTION_START)
            and "ENVELOPE" not in upper
        ):
            in_disp_section = True
            current_case_header = None
            continue

        if not in_disp_section:
            continue

        if any(keyword in upper for keyword in DISP_SECTION_END):
            in_disp_section = False
            current_case_header = None
            continue

        if line.strip().startswith("--") and "DISPLACEMENTS" not in upper:
            in_disp_section = False
            current_case_header = None
            continue

        case_match = re.search(r"CASE\s+(\d+):(.*)", upper)
        if case_match:
            current_case_header = case_match.group(1)
            if "OPERATIONAL" in case_match.group(2):
                service_cases.add(current_case_header)
            continue

        tokens = line.strip().split()
        case_id = x_val = y_val = None

        if len(tokens) >= 8 and tokens[0].isdigit() and tokens[1].isdigit():
            case_id, x_val, y_val = tokens[1], tokens[2], tokens[3]
        elif len(tokens) == 7 and tokens[0].isdigit() and not current_case_header:
            case_id, x_val, y_val = tokens[0], tokens[1], tokens[2]
        elif current_case_header and len(tokens) >= 7 and tokens[0].isdigit():
            case_id, x_val, y_val = current_case_header, tokens[1], tokens[2]

        if not (case_id and x_val and y_val):
            continue

        if service_cases and case_id not in service_cases:
            continue

        x_disp = num(x_val)
        y_disp = num(y_val)
        if x_disp is None or y_disp is None:
            continue

        # Keep the original engineering rule: governing horizontal
        # displacement = max(abs(X), abs(Y)).
        max_xy = max(abs(x_disp), abs(y_disp))
        if max_xy > max_disp and max_xy < 15000:
            max_disp = max_xy
            case_disp = case_id

    # -----------------------------------------------------
    # 2A. ENVELOPE OF TOWER ROTATIONS
    # -----------------------------------------------------
    in_rot_env = False
    max_sway_env, case_sway_env = 0.0, ""
    max_twist_env, case_twist_env = 0.0, ""
    env_found = False

    for line in lines:
        if re.search(r"ENVELOPE\s+OF\s+TOWER\s+ROTATIONS", line, re.I):
            in_rot_env = True
            continue

        if not in_rot_env:
            continue

        if not line.strip() or re.search(r"Pt\s+Height", line, re.I):
            continue

        if (
            (
                re.search(r"ENVELOPE", line, re.I)
                and not re.search(r"ENVELOPE\s+OF\s+TOWER\s+ROTATIONS", line, re.I)
            )
            or "--" in line
            or "M E M B E R" in line.upper()
        ):
            in_rot_env = False
            continue

        tokens = line.strip().split()
        if len(tokens) < 8 or not tokens[0].isdigit():
            continue

        try:
            lc_x, x_rot = tokens[2], abs(float(tokens[3]))
            lc_y, y_rot = tokens[4], abs(float(tokens[5]))
            lc_z, z_rot = tokens[6], abs(float(tokens[7]))
        except ValueError:
            continue

        env_found = True

        if not service_cases or lc_x in service_cases:
            if x_rot > max_sway_env:
                max_sway_env, case_sway_env = x_rot, lc_x

        if not service_cases or lc_y in service_cases:
            if y_rot > max_sway_env:
                max_sway_env, case_sway_env = y_rot, lc_y

        if not service_cases or lc_z in service_cases:
            if z_rot > max_twist_env:
                max_twist_env, case_twist_env = z_rot, lc_z

    max_sway = max_sway_env if env_found else 0.0
    case_sway = case_sway_env if env_found else ""
    max_twist = max_twist_env if env_found else 0.0
    case_twist = case_twist_env if env_found else ""

    # -----------------------------------------------------
    # 3. LATTICE SR
    # -----------------------------------------------------
    sr_data: List[List[str]] = []
    capturing_sr = False

    for line in lines:
        if "Pnl" in line and "Members" in line:
            capturing_sr = True
            continue

        if not capturing_sr:
            continue

        if not line.strip():
            continue

        if "Mass Summary" in line or "SUPPORT REACTIONS" in line:
            if sr_data:
                break
            continue

        line_fixed = re.sub(r"(\d+)\s*-\s*(\d+)", r"\1-\2", line)
        tokens = line_fixed.strip().split()
        if tokens and tokens[0].isdigit():
            sr_data.append(tokens)

    # -----------------------------------------------------
    # 3A. POLE SR
    # -----------------------------------------------------
    pole_sr_data: List[List[str]] = []
    capturing_pole_sr = False

    for line in lines:
        low = line.lower()

        if "memb" in low and "ht" in low and "ratio" in low:
            capturing_pole_sr = True
            continue

        if not capturing_pole_sr:
            continue

        if not line.strip():
            continue

        if any(
            kw in line
            for kw in (
                "Mass Summary",
                "SUPPORT REACTIONS",
                "NODE DISPLACEMENTS",
                "ENVELOPE",
                "Page",
            )
        ):
            capturing_pole_sr = False
            continue

        tokens = line.strip().split()
        if len(tokens) >= 15 and tokens[0].isdigit():
            if num(tokens[1]) is not None:
                pole_sr_data.append(tokens[:16])

    # -----------------------------------------------------
    # 4. MASS SUMMARY
    # -----------------------------------------------------
    mass_data: List[List[str]] = []
    mass_pattern = re.search(
        r"Mass Summary\s*\n\s*Sect Size.*?M \(kg\)\n(.*?)\s+-{5,}",
        content,
        re.DOTALL,
    )
    if mass_pattern:
        for line in mass_pattern.group(1).strip().splitlines():
            if line.strip():
                mass_data.append(re.split(r"\s{2,}", line.strip()))

    # -----------------------------------------------------
    # 5. SUPPORT REACTIONS
    # -----------------------------------------------------
    reactions_data: List[List[str]] = []
    reaction_pattern = re.search(
        r"SUPPORT REACTIONS \(Applied to tower\)\s*\n\s*Case\s+Node\s+FX.*?\n"
        r"(.*?)(?:\n\s*\n\s*\n|\Z)",
        content,
        re.DOTALL,
    )

    if reaction_pattern:
        current_case_reaction = ""
        for line in reaction_pattern.group(1).strip().splitlines():
            if not line.strip() or "Resultant" in line:
                continue

            cols = re.split(r"\s{2,}", line.strip())

            if len(cols) >= 8:
                current_case_reaction = cols[0]
                reactions_data.append(cols)
            elif len(cols) == 7:
                cols.insert(0, current_case_reaction)
                reactions_data.append(cols)
            elif cols and cols[0]:
                if re.match(r"^\d+$", cols[0]):
                    current_case_reaction = cols[0]
                    reactions_data.append(cols)
                else:
                    cols.insert(0, current_case_reaction)
                    reactions_data.append(cols)

    # -----------------------------------------------------
    # BUILD DATAFRAMES
    # -----------------------------------------------------
    tower_type = "Unknown"
    total_height = 0.0

    if sr_data:
        tower_type = "Lattice Tower"
        sr_headers = [
            "Elev (m)", "Pnl", "Members", "Typ", "Size", "fy", "nb",
            "Comp_Case", "Comp_Pu", "Comp_Curve", "Comp_KL/r", "Comp_Dir",
            "Comp_Pn", "Comp_Pu/Pn", "Tens_Case", "Tens_Pu", "Tens_Pn",
            "Tens_Pu/Pn", "Joint_Case", "Joint_Pu", "Joint_Dia", "Joint_Grade",
            "Joint_Type", "Joint_Pn", "Joint_Pu/Pn"
        ]

        normalized_sr = []
        for row in sr_data:
            row = list(row)
            pnl_val = row[0] if row else ""
            elev_val = cumulative_elev.get(pnl_val, "")

            if len(row) > 9:
                comp_val, comp_dir = split_klr(row[9])
                row[9] = comp_val
                row.insert(10, comp_dir)

            row = [elev_val] + row
            if len(row) < len(sr_headers):
                row += [""] * (len(sr_headers) - len(row))
            else:
                row = row[: len(sr_headers)]
            normalized_sr.append(row)

        df_sr = pd.DataFrame(normalized_sr, columns=sr_headers)
        dfs["Data SR"] = df_sr

        temp = df_sr.copy()
        for col in ("Comp_Pu/Pn", "Tens_Pu/Pn", "Comp_KL/r"):
            temp[col] = pd.to_numeric(temp[col], errors="coerce")

        resume_data = []
        for member_type in temp["Typ"].dropna().unique():
            if str(member_type).strip() == "":
                continue

            subset = temp[temp["Typ"] == member_type]

            def max_info(frame, value_col, dir_col=None):
                valid = frame.dropna(subset=[value_col])
                if valid.empty:
                    return (None, None, None, None) if dir_col else (None, None, None)

                idx = valid[value_col].idxmax()
                if dir_col:
                    return (
                        valid.loc[idx, value_col],
                        valid.loc[idx, dir_col],
                        valid.loc[idx, "Pnl"],
                        valid.loc[idx, "Elev (m)"],
                    )
                return (
                    valid.loc[idx, value_col],
                    valid.loc[idx, "Pnl"],
                    valid.loc[idx, "Elev (m)"],
                )

            max_comp, pnl_comp, elev_comp = max_info(subset, "Comp_Pu/Pn")
            max_tens, pnl_tens, elev_tens = max_info(subset, "Tens_Pu/Pn")
            max_klr, dir_klr, pnl_klr, elev_klr = max_info(
                subset, "Comp_KL/r", "Comp_Dir"
            )

            vc = max_comp if pd.notna(max_comp) else -1
            vt = max_tens if pd.notna(max_tens) else -1

            if vc == -1 and vt == -1:
                max_ratio, elev_max, pnl_max, remark = None, None, None, ""
            elif vc >= vt:
                max_ratio, elev_max, pnl_max, remark = max_comp, elev_comp, pnl_comp, "compression"
            else:
                max_ratio, elev_max, pnl_max, remark = max_tens, elev_tens, pnl_tens, "tension"

            resume_data.append(
                {
                    "Typ": member_type,
                    "Max Comp_Pu/Pn": max_comp,
                    "Elev (Comp)": elev_comp,
                    "Pnl (Comp)": pnl_comp,
                    "Max Tens_Pu/Pn": max_tens,
                    "Elev (Tens)": elev_tens,
                    "Pnl (Tens)": pnl_tens,
                    "Max Pu/Pn": max_ratio,
                    "Elev (Max)": elev_max,
                    "Pnl (Max)": pnl_max,
                    "Remark": remark,
                    "Max Comp_KL/r": max_klr,
                    "Dir (C_KL/r)": dir_klr,
                    "Elev (C_KL/r)": elev_klr,
                    "Pnl (C_KL/r)": pnl_klr,
                }
            )

        if resume_data:
            dfs["Resume"] = pd.DataFrame(resume_data)

    if pole_sr_data:
        tower_type = "Monopole" if not sr_data else "Hybrid (Pole+Lattice)"

        pole_headers = [
            "Memb", "Ht (m)", "D (mm)", "t (mm)", "fy", "fy_eff", "LC",
            "Nu (kN)", "Mxu (kNm)", "Myu (kNm)", "Mru (kNm)", "Tu (kNm)",
            "phi_Nn", "phi_Mn", "phi_Tn", "Ratio"
        ]

        normalized_pole = []
        for row in pole_sr_data:
            row = list(row)
            if len(row) < 16:
                row += [""] * (16 - len(row))
            else:
                row = row[:16]
            normalized_pole.append(row)

        df_pole = pd.DataFrame(normalized_pole, columns=pole_headers)
        dfs["Pole SR"] = df_pole

        temp = df_pole.copy()
        temp["Ratio"] = pd.to_numeric(temp["Ratio"], errors="coerce")
        temp["Ht (m)"] = pd.to_numeric(temp["Ht (m)"], errors="coerce")

        if temp["Ratio"].notna().any():
            idx = temp["Ratio"].idxmax()
            dfs["Pole Resume"] = pd.DataFrame(
                [
                    {
                        "Structure Type": "Cantilevered Pole",
                        "Max Ratio": temp.loc[idx, "Ratio"],
                        "Elev / Ht (m)": temp.loc[idx, "Ht (m)"],
                        "Member ID": temp.loc[idx, "Memb"],
                        "Load Case (LC)": temp.loc[idx, "LC"],
                    }
                ]
            )

    # -----------------------------------------------------
    # 6. OVERALL CHECK
    # -----------------------------------------------------
    if sr_data or pole_sr_data:
        is_pole_only = bool(pole_sr_data) and not bool(sr_data)

        # Preserved from current application rules:
        # pole H/100, lattice H/200.
        disp_divisor = 100.0 if is_pole_only else 200.0
        rot_limit = 1.0 if is_pole_only else 0.5

        if cumulative_elev:
            total_height = max(cumulative_elev.values())

        if total_height == 0.0 and pole_sr_data:
            heights = [num(row[1]) for row in pole_sr_data if len(row) > 1]
            heights = [h for h in heights if h is not None]
            if heights:
                total_height = max(heights)

        # Displacement is displayed and exported in metres.
        # MS Tower node displacement values are already parsed in metres.
        limit_disp_m = total_height / disp_divisor if total_height > 0 else 0.0
        max_disp_m = max_disp

        overall_data = {
            "Item": [
                "Displacement",
                "Twist",
                "Sway",
                "Anchor Bolt",
            ],
            "Max Value": [
                round(max_disp_m, 4) if max_disp > 0 else None,
                round(max_twist, 4) if max_twist > 0 else None,
                round(max_sway, 4) if max_sway > 0 else None,
                None,
            ],
            "Unit": ["m", "deg", "deg", "-",],
            "Allowable Limit": [
                round(limit_disp_m, 4) if limit_disp_m > 0 else None,
                rot_limit,
                rot_limit,
                1.0,
            ],
            "Limit Basis": [
                f"H/{int(disp_divisor)}",
                "Tower rotation limit",
                "Tower rotation limit",
                "Fixed check value",
            ],
            "Remark (Load Case)": [
                f"CASE {case_disp}" if case_disp else "",
                f"CASE {case_twist}" if case_twist else "",
                f"CASE {case_sway}" if case_sway else "",
                "",
            ],
            "Status": [
                status_label(max_disp_m, limit_disp_m),
                status_label(max_twist, rot_limit),
                status_label(max_sway, rot_limit),
                "N/A",
            ],
        }
        dfs["Overall Check"] = pd.DataFrame(overall_data)

    # -----------------------------------------------------
    # 7. MASS
    # -----------------------------------------------------
    if mass_data:
        headers = ["Sect", "Size", "fy", "L (m)", "M (kg)"]
        normalized = []
        for row in mass_data:
            row = list(row)
            if len(row) < len(headers):
                row += [""] * (len(headers) - len(row))
            normalized.append(row[: len(headers)])
        dfs["Mass Summary"] = pd.DataFrame(normalized, columns=headers)

    # -----------------------------------------------------
    # 8. REACTIONS + GOVERNING ABSOLUTE
    # -----------------------------------------------------
    if reactions_data:
        reaction_headers = [
            "Case", "Node", "FX", "FY", "FZ", "MX", "MY", "MZ", "Note"
        ]
        normalized_reactions = []

        for row in reactions_data:
            row = list(row)
            if len(row) < len(reaction_headers):
                row += [""] * (len(reaction_headers) - len(row))
            elif len(row) > len(reaction_headers):
                row = (
                    row[: len(reaction_headers) - 1]
                    + [" ".join(row[len(reaction_headers) - 1:])]
                )
            normalized_reactions.append(row)

        df_reactions = pd.DataFrame(normalized_reactions, columns=reaction_headers)
        for col in ("FX", "FY", "FZ", "MX", "MY", "MZ"):
            df_reactions[col] = pd.to_numeric(df_reactions[col], errors="coerce")

        dfs["Support Reactions"] = df_reactions

        reaction_resume_list = []
        for col in ("FX", "FY", "FZ", "MX", "MY", "MZ"):
            valid = df_reactions.dropna(subset=[col])
            if valid.empty:
                continue

            idx_max = valid[col].idxmax()
            idx_min = valid[col].idxmin()
            idx_abs = valid[col].abs().idxmax()

            reaction_resume_list.append(
                {
                    "Force/Moment": col,
                    "Max Positive (+)": valid.loc[idx_max, col],
                    "Node (+)": valid.loc[idx_max, "Node"],
                    "Case (+)": valid.loc[idx_max, "Case"],
                    "Max Negative (-)": valid.loc[idx_min, col],
                    "Node (-)": valid.loc[idx_min, "Node"],
                    "Case (-)": valid.loc[idx_min, "Case"],
                    "Governing Abs": valid.loc[idx_abs, col],
                    "Node (Abs)": valid.loc[idx_abs, "Node"],
                    "Case (Abs)": valid.loc[idx_abs, "Case"],
                }
            )

        if reaction_resume_list:
            dfs["Reaction Resume"] = pd.DataFrame(reaction_resume_list)

    # -----------------------------------------------------
    # 9. LOAD PARAMETERS + TOWER LOADING
    # -----------------------------------------------------
    load_parameters = extract_load_parameters(content)
    antenna_loading = extract_antenna_loading(content)
    antenna_height_resume = build_antenna_height_resume(antenna_loading)
    linear_loading = extract_linear_loading(content)

    if not load_parameters.empty:
        dfs["Load Parameters"] = load_parameters
    if not antenna_height_resume.empty:
        dfs["Antenna Height Resume"] = antenna_height_resume
    if not antenna_loading.empty:
        dfs["Antenna Loading"] = antenna_loading
    if not linear_loading.empty:
        dfs["Linear Loading"] = linear_loading

    # -----------------------------------------------------
    # 10. TOWER MEMBER DESCRIPTION + BOLT DATA
    # -----------------------------------------------------
    member_section_catalog = extract_member_section_catalog(content)
    tower_member_detail, tower_member_resume, tower_profile = extract_tower_member_data(
        content, member_section_catalog
    )
    bolt_data = extract_bolt_data(content)

    if not tower_profile.empty:
        dfs["Tower Profile"] = tower_profile
    if not tower_member_resume.empty:
        dfs["Tower Member Resume"] = tower_member_resume
    if not tower_member_detail.empty:
        dfs["Tower Member Detail"] = tower_member_detail
    if not member_section_catalog.empty:
        dfs["Member Section Catalog"] = member_section_catalog
    if not bolt_data.empty:
        dfs["Bolt Data"] = bolt_data

    # -----------------------------------------------------
    # METADATA / VALIDATION
    # -----------------------------------------------------
    if total_height <= 0:
        warnings.append("Total height could not be determined.")
    if not sr_data and not pole_sr_data:
        warnings.append("No lattice/pole member strength data detected.")
    if max_disp == 0:
        warnings.append("Horizontal node displacement was not detected.")
    if not env_found:
        warnings.append("Envelope of tower rotations was not detected.")
    if not reactions_data:
        warnings.append("Support reactions section was not detected.")

    metadata = {
        "Tower Type": tower_type,
        "Total Height (m)": round(total_height, 2) if total_height else None,
        "Service Cases": ", ".join(sorted(service_cases)) if service_cases else "",
        "Max Displacement (m)": round(max_disp, 4) if max_disp else None,
        "Max Sway (deg)": round(max_sway, 4) if max_sway else None,
        "Max Twist (deg)": round(max_twist, 4) if max_twist else None,
        "Warnings": warnings,
    }

    return dfs, metadata


@st.cache_data(show_spinner=False)
def generate_excel_cached(dfs: Dict[str, pd.DataFrame]) -> bytes:
    return df_to_excel_bytes(dfs)


# =========================================================
# ENGINEERING SUMMARY
# =========================================================
def build_file_summary(
    filename: str, dfs: Dict[str, pd.DataFrame], metadata: Dict[str, Any]
) -> Dict[str, Any]:
    result = {
        "Filename": filename,
        "Tower Type": metadata.get("Tower Type", "Unknown"),
        "Height (m)": metadata.get("Total Height (m)"),
        "Max Ratio": None,
        "Disp (m)": metadata.get("Max Displacement (m)"),
        "Sway (deg)": metadata.get("Max Sway (deg)"),
        "Twist (deg)": metadata.get("Max Twist (deg)"),
        "Overall": "N/A",
        "Status": "⚠️ Review",
    }

    ratio_candidates = []

    if "Resume" in dfs and "Max Pu/Pn" in dfs["Resume"].columns:
        ratio_candidates += pd.to_numeric(
            dfs["Resume"]["Max Pu/Pn"], errors="coerce"
        ).dropna().tolist()

    if "Pole Resume" in dfs and "Max Ratio" in dfs["Pole Resume"].columns:
        ratio_candidates += pd.to_numeric(
            dfs["Pole Resume"]["Max Ratio"], errors="coerce"
        ).dropna().tolist()

    if ratio_candidates:
        result["Max Ratio"] = max(ratio_candidates)

    overall_pass = []
    if "Overall Check" in dfs:
        oc = dfs["Overall Check"]
        if "Status" in oc.columns:
            vals = oc["Status"].astype(str).tolist()
            if "FAIL" in vals:
                result["Overall"] = "FAIL"
            elif any(v == "PASS" for v in vals):
                result["Overall"] = "PASS"
            else:
                result["Overall"] = "N/A"

    # Member ratio is part of overall screening, not a substitute for the
    # structural report's complete design code checks.
    if result["Max Ratio"] is not None:
        overall_pass.append(result["Max Ratio"] <= 1.0)

    if result["Overall"] == "FAIL":
        result["Status"] = "🔴 FAIL"
    elif overall_pass and not all(overall_pass):
        result["Status"] = "🔴 FAIL"
    elif result["Overall"] == "PASS":
        result["Status"] = "🟢 PASS"
    else:
        result["Status"] = "🟠 REVIEW"

    return result


def overall_counts(summary_df: pd.DataFrame):
    if summary_df.empty:
        return 0, 0, 0
    pass_n = int((summary_df["Status"] == "🟢 PASS").sum())
    fail_n = int((summary_df["Status"] == "🔴 FAIL").sum())
    review_n = len(summary_df) - pass_n - fail_n
    return pass_n, fail_n, review_n


# =========================================================
# UI COMPONENTS
# =========================================================
def inject_css():
    st.markdown(
        """
        <style>
        .block-container {padding-top: 1rem; padding-bottom: 2rem;}
        .metric-card {
            border: 1px solid rgba(128,128,128,.25);
            border-radius: 12px;
            padding: 12px 14px;
            background: rgba(128,128,128,.045);
        }
        .section-title {
            margin-top: .4rem;
            margin-bottom: .2rem;
        }
        div[data-testid="stDataFrame"] {
            border-radius: 10px;
            overflow: hidden;
        }
        .small-muted {
            color: rgba(128,128,128,.95);
            font-size: 0.85rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_kpis(summary_df: pd.DataFrame):
    pass_n, fail_n, review_n = overall_counts(summary_df)
    max_ratio = (
        pd.to_numeric(summary_df["Max Ratio"], errors="coerce").max()
        if not summary_df.empty
        else None
    )
    max_disp = (
        pd.to_numeric(summary_df["Disp (m)"], errors="coerce").max()
        if not summary_df.empty
        else None
    )

    cols = st.columns(5)
    cols[0].metric("Files", len(summary_df))
    cols[1].metric("PASS", pass_n)
    cols[2].metric("FAIL", fail_n)
    cols[3].metric("REVIEW", review_n)
    cols[4].metric("Highest Pu/Pn", f"{max_ratio:.3f}" if pd.notna(max_ratio) else "—")

    st.caption(
        f"Highest displacement detected: "
        f"{max_disp:.4f} m" if pd.notna(max_disp) else
        "Highest displacement detected: —"
    )


def render_overall_check(df: pd.DataFrame):
    if df.empty:
        return

    show = df.copy()
    st.dataframe(
        show,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Status": st.column_config.TextColumn("Status", width="small"),
            "Max Value": st.column_config.NumberColumn("Max Value", format="%.4f"),
            "Allowable Limit": st.column_config.NumberColumn(
                "Allowable", format="%.4f"
            ),
        },
    )

    failed = show[show["Status"].astype(str) == "FAIL"]
    if not failed.empty:
        st.error(
            "Governing serviceability check failed. Review the corresponding load case "
            "and the full structural report before approval."
        )


def render_member_resume(df: pd.DataFrame):
    if df.empty:
        return

    cols = ["Typ", "Max Pu/Pn", "Elev (Max)", "Pnl (Max)", "Remark"]
    cols = [c for c in cols if c in df.columns]

    out = df[cols].copy()
    if "Max Pu/Pn" in out.columns:
        out["Max Pu/Pn"] = pd.to_numeric(out["Max Pu/Pn"], errors="coerce")
        out = out.sort_values("Max Pu/Pn", ascending=False, na_position="last")

    st.dataframe(
        out,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Max Pu/Pn": st.column_config.NumberColumn("Max Pu/Pn", format="%.3f")
        },
    )


def render_governing_metrics(dfs: Dict[str, pd.DataFrame], metadata: Dict[str, Any]):
    st.markdown("### Governing engineering values")

    vals = st.columns(5)
    vals[0].metric("Tower Type", metadata.get("Tower Type", "Unknown"))
    vals[1].metric(
        "Height",
        f"{metadata['Total Height (m)']:.2f} m"
        if metadata.get("Total Height (m)") is not None
        else "—",
    )

    max_ratio = None
    governing_member = "—"

    if "Resume" in dfs and "Max Pu/Pn" in dfs["Resume"].columns:
        s = pd.to_numeric(dfs["Resume"]["Max Pu/Pn"], errors="coerce")
        if s.notna().any():
            idx = s.idxmax()
            max_ratio = float(s.loc[idx])
            governing_member = str(dfs["Resume"].loc[idx, "Typ"])

    if "Pole Resume" in dfs and "Max Ratio" in dfs["Pole Resume"].columns:
        s = pd.to_numeric(dfs["Pole Resume"]["Max Ratio"], errors="coerce")
        if s.notna().any() and (max_ratio is None or float(s.max()) > max_ratio):
            idx = s.idxmax()
            max_ratio = float(s.loc[idx])
            governing_member = str(dfs["Pole Resume"].loc[idx, "Member ID"])

    vals[2].metric("Governing Ratio", f"{max_ratio:.3f}" if max_ratio is not None else "—")
    vals[3].metric("Governing Member", governing_member)
    vals[4].metric(
        "Displacement",
        f"{metadata['Max Displacement (m)']:.4f} m"
        if metadata.get("Max Displacement (m)") is not None
        else "—",
    )


# =========================================================
# MAIN
# =========================================================
inject_css()

st.title(APP_TITLE)
st.caption(APP_SUBTITLE)

with st.expander("What this version improves", expanded=False):
    st.markdown(
        """
        **Performance:** parsing is cached, Excel is generated lazily, and the inspector
        uses one selected file instead of rendering dozens of tabs at once.

        **Engineering UX:** the first screen focuses on governing ratio, displacement,
        sway/twist, height, status, and load-case traceability.

        **Batch workflow:** one summary table for all towers, filtering, sorting,
        batch ZIP export, and individual Excel export.

        **Data quality:** missing sections become explicit warnings instead of silently
        looking like a clean result.
        """
    )

uploaded_files = st.file_uploader(
    "Upload MS Tower reports",
    type=["rpt", "txt", "mst", "udp"],
    accept_multiple_files=True,
    help=(
        "Upload one or many .rpt/.txt/.mst report files. Optional .udp files may "
        "be uploaded in the same control and are used only as custom geometry "
        "components in Tab 3."
    ),
)

if not uploaded_files:
    st.info(
        "Upload one or more MS Tower report files to start. "
        "Optional UDP files can be added together with the reports for custom drawing."
    )
    st.stop()

# -------------------------------------------------------------------------
# Upload role detection
# -------------------------------------------------------------------------
# A file can legitimately be BOTH a parent MStower/report source and a UDP
# source.  This happens with text exports/big-pastes that contain PROFILE plus
# one or more embedded UDP NODE/MEMB definitions.  Do not remove such files
# from report_files merely because UDP syntax is present.
def _looks_like_udp_text(text: str) -> bool:
    upper = str(text).upper().replace("\r", "\n")
    return bool(
        re.search(r"(?:^|\n)\s*UDP\s+[@A-Z0-9_.-]+", upper)
        and re.search(r"(?:^|\n)\s*NODE\s+\d+", upper)
        and re.search(r"(?:^|\n)\s*MEMB\s+\d+", upper)
    )


def _looks_like_parent_report_text(text: str) -> bool:
    """Identify a parent MST/report independently of embedded UDP content."""
    upper = str(text).upper().replace("\r", "\n")

    # PROFILE is the strongest geometry-parent marker.  PARAMETERS / analysis
    # report headings cover .rpt/.txt exports that may not contain PROFILE.
    strong_markers = (
        r"(?:^|\n)\s*PROFILE\b",
        r"(?:^|\n)\s*PARAMETERS\b",
        r"NODE\s+DISPLACEMENTS",
        r"ENVELOPE\s+OF\s+TOWER\s+ROTATIONS",
        r"SUPPORT\s+REACTIONS",
        r"(?:^|\n)\s*ANCILLARIES\b",
    )
    if any(re.search(pattern, upper) for pattern in strong_markers):
        return True

    # A normal model text generally carries title + sections and/or panel
    # declarations.  A pure UDP file can also have SECTIONS, so SECTIONS alone
    # must never classify it as a parent report.
    has_title = bool(re.search(r"(?:^|\n)\s*TITL[12]\b", upper))
    has_panel = bool(re.search(r"(?:^|\n)\s*PANEL\s+\d+\s+HT\b", upper))
    has_sections = bool(re.search(r"(?:^|\n)\s*SECTIONS\b", upper))
    return has_panel or (has_title and has_sections and "PROFILE" in upper)


_decoded_uploads: Dict[str, str] = {
    f.name: decode_rpt(f.getvalue()) for f in uploaded_files
}

report_files = []
udp_uploaded_files = []

for f in uploaded_files:
    name_lower = str(f.name).lower()
    text = _decoded_uploads[f.name]
    ext = name_lower.rsplit(".", 1)[-1] if "." in name_lower else ""
    has_udp = _looks_like_udp_text(text)
    has_parent = _looks_like_parent_report_text(text)

    # Explicit .udp is a UDP source.  Any other uploaded file that embeds UDP
    # syntax is ALSO made available to the UDP resolver.
    if ext == "udp" or has_udp:
        udp_uploaded_files.append(f)

    # .rpt/.mst are always parent/report candidates even if they embed UDP.
    # .txt remains a report when it contains parent markers; an ambiguous text
    # that is not a pure UDP is also kept as a report for backward compatibility.
    if ext in {"rpt", "mst"}:
        report_files.append(f)
    elif ext == "txt":
        if has_parent or not has_udp:
            report_files.append(f)
    elif ext != "udp" and has_parent:
        report_files.append(f)

udp_sources: Dict[str, str] = {
    f.name: _decoded_uploads[f.name] for f in udp_uploaded_files
}

if not report_files:
    st.warning(
        "No parent MS Tower report/model was detected. The uploaded file(s) look "
        "like standalone UDP geometry. Upload the parent .mst/.rpt/.txt model as "
        "well; files that contain both PROFILE and UDP data are now accepted as "
        "both report and UDP source."
    )
    st.stop()

# ---------------------------------------------------------
# Parse batch — no Excel generation here
# ---------------------------------------------------------
results: List[Dict[str, Any]] = []

progress = st.progress(0, text="Preparing batch...")
for i, file in enumerate(report_files, start=1):
    try:
        content = decode_rpt(file.getvalue())
        dfs, metadata = parse_rpt_cached(content)

        summary = build_file_summary(file.name, dfs, metadata)

        results.append(
            {
                "Filename": file.name,
                "Tower Type": metadata.get("Tower Type", "Unknown"),
                "Height (m)": metadata.get("Total Height (m)"),
                "Max Ratio": summary["Max Ratio"],
                "Disp (m)": metadata.get("Max Displacement (m)"),
                "Sway (deg)": metadata.get("Max Sway (deg)"),
                "Twist (deg)": metadata.get("Max Twist (deg)"),
                "Overall": summary["Overall"],
                "Status": summary["Status"],
                "dfs": dfs,
                "metadata": metadata,
                "content": content,
            }
        )
    except Exception as exc:
        results.append(
            {
                "Filename": file.name,
                "Tower Type": "Unknown",
                "Height (m)": None,
                "Max Ratio": None,
                "Disp (m)": None,
                "Sway (deg)": None,
                "Twist (deg)": None,
                "Overall": "ERROR",
                "Status": "❌ Error",
                "dfs": {},
                "metadata": {"Warnings": [str(exc)]},
                "content": content if "content" in locals() else "",
            }
        )

    progress.progress(
        i / len(report_files),
        text=f"Parsed {i}/{len(report_files)} report file(s)",
    )

progress.empty()

summary_df = pd.DataFrame(
    [
        {k: row[k] for k in row.keys() if k not in ("dfs", "metadata", "content")}
        for row in results
    ]
)

# ---------------------------------------------------------
# Dashboard
# ---------------------------------------------------------
st.divider()
st.subheader("📊 Batch Engineering Dashboard")
render_kpis(summary_df)

c1, c2 = st.columns([1.3, 1.0])
with c1:
    status_filter = st.multiselect(
        "Status filter",
        options=sorted(summary_df["Status"].dropna().unique().tolist()),
        default=[],
    )
with c2:
    type_filter = st.multiselect(
        "Tower type filter",
        options=sorted(summary_df["Tower Type"].dropna().unique().tolist()),
        default=[],
    )

view_df = summary_df.copy()
if status_filter:
    view_df = view_df[view_df["Status"].isin(status_filter)]
if type_filter:
    view_df = view_df[view_df["Tower Type"].isin(type_filter)]

if "Max Ratio" in view_df.columns:
    view_df = view_df.sort_values(
        by="Max Ratio", ascending=False, na_position="last"
    )

st.dataframe(
    view_df,
    use_container_width=True,
    hide_index=True,
    column_config={
        "Filename": st.column_config.TextColumn("File", width="large"),
        "Height (m)": st.column_config.NumberColumn("Height", format="%.2f"),
        "Max Ratio": st.column_config.NumberColumn("Max Pu/Pn / Ratio", format="%.3f"),
        "Disp (m)": st.column_config.NumberColumn("Disp (m)", format="%.4f"),
        "Sway (deg)": st.column_config.NumberColumn("Sway", format="%.4f"),
        "Twist (deg)": st.column_config.NumberColumn("Twist", format="%.4f"),
        "Overall": st.column_config.TextColumn("Overall", width="small"),
        "Status": st.column_config.TextColumn("Review Status", width="small"),
    },
)

# ---------------------------------------------------------
# Export area
# ---------------------------------------------------------
st.subheader("📦 Export")

valid_results = [r for r in results if r["dfs"]]

if valid_results:
    if st.button(
        f"Build ZIP of {len(valid_results)} Excel report(s)",
        type="primary",
        use_container_width=True,
    ):
        with st.spinner("Generating Excel files and ZIP..."):
            zip_buffer = io.BytesIO()
            used_names = set()

            with zipfile.ZipFile(
                zip_buffer, "w", compression=zipfile.ZIP_DEFLATED
            ) as zf:
                for item in valid_results:
                    excel_bytes = generate_excel_cached(item["dfs"])
                    stem = item["Filename"].rsplit(".", 1)[0]
                    name = f"{stem}.xlsx"

                    # Avoid duplicate names inside ZIP.
                    if name in used_names:
                        base, ext = name.rsplit(".", 1)
                        n = 2
                        while f"{base}_{n}.{ext}" in used_names:
                            n += 1
                        name = f"{base}_{n}.{ext}"

                    used_names.add(name)
                    zf.writestr(name, excel_bytes)

            zip_bytes = zip_buffer.getvalue()

        st.download_button(
            "⬇️ Download ZIP",
            data=zip_bytes,
            file_name="MS_Tower_Extracted_Batch_Pro.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True,
        )

# ---------------------------------------------------------
# File inspector — ONE file at a time
# ---------------------------------------------------------
st.divider()
st.subheader("🔎 File Inspector")

file_options = [r["Filename"] for r in results]

# Keep the selector compact and place the individual Excel download beside it.
select_col, download_col = st.columns([3.4, 1.0])

with select_col:
    selected_name = st.selectbox(
        "Select report to inspect",
        options=file_options,
        index=0,
    )

selected = next(r for r in results if r["Filename"] == selected_name)
dfs = selected["dfs"]
metadata = selected["metadata"]
selected_content = selected.get("content", "")

with download_col:
    # Align the button with the selectbox control (below its label).
    st.markdown("<div style='height: 28px'></div>", unsafe_allow_html=True)
    if dfs:
        selected_excel = generate_excel_cached(dfs)
        st.download_button(
            "⬇️ Download Excel",
            data=selected_excel,
            file_name=f"{selected_name.rsplit('.', 1)[0]}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"download_{selected_name}",
        )
    else:
        st.button(
            "Excel unavailable",
            disabled=True,
            use_container_width=True,
            key=f"download_disabled_{selected_name}",
        )

report_tab, load_tab, member_tab = st.tabs(
    [
        "📊 Report Hasil Analisa",
        "⚙️ Load Parameter",
        "🗼 Keterangan Member Tower",
    ]
)

# =========================================================
# TAB 1 — REPORT HASIL ANALISA
# =========================================================
with report_tab:
    render_governing_metrics(dfs, metadata)

    left, right = st.columns([1.15, 1.0])

    with left:
        st.markdown("### Decision summary")

        if "Overall Check" in dfs:
            render_overall_check(dfs["Overall Check"])

        if "Resume" in dfs:
            st.markdown("### Governing lattice members")
            render_member_resume(dfs["Resume"])

        if "Pole Resume" in dfs:
            st.markdown("### Governing pole")
            st.dataframe(
                dfs["Pole Resume"],
                use_container_width=True,
                hide_index=True,
            )

    with right:
        # Load-related datasets are shown in Tab 2 to keep the analysis tab focused.
        load_sheet_names = {"Load Parameters", "Antenna Height Resume", "Antenna Loading", "Linear Loading"}
        member_sheet_names = {
            "Tower Profile", "Tower Member Resume", "Tower Member Detail",
            "Member Section Catalog", "Bolt Data",
        }
        sheet_options = [
            name for name in dfs.keys()
            if name not in load_sheet_names and name not in member_sheet_names
        ]

        raw_title_col, raw_select_col = st.columns(
            [1.15, 1.0], vertical_alignment="center"
        )

        with raw_title_col:
            st.markdown("### Raw extracted data")

        if sheet_options:
            with raw_select_col:
                selected_sheet = st.selectbox(
                    "Sheet / dataset",
                    options=sheet_options,
                    key=f"sheet_{selected_name}",
                    label_visibility="collapsed",
                )

            selected_df = dfs[selected_sheet]

            overall_rows = len(dfs.get("Overall Check", pd.DataFrame()))
            resume_rows = len(dfs.get("Resume", pd.DataFrame()))
            pole_rows = len(dfs.get("Pole Resume", pd.DataFrame()))

            left_visual_height = 0
            if overall_rows:
                left_visual_height += 50 + (overall_rows * 35)
            if resume_rows:
                left_visual_height += 88 + 35 + (resume_rows * 35)
            if pole_rows:
                left_visual_height += 88 + 35 + (pole_rows * 35)

            raw_table_height = max(430, min(left_visual_height, 720))

            st.dataframe(
                selected_df,
                use_container_width=True,
                hide_index=True,
                height=raw_table_height,
            )
            st.caption(
                f"{selected_sheet}: {len(selected_df):,} row(s) × "
                f"{len(selected_df.columns):,} column(s)"
            )
        else:
            st.error(
                "No extractable engineering data was detected from this report. "
                "Check report format/version or inspect the source .rpt."
            )

    # Preserve traceability without consuming permanent dashboard space.
    with st.expander("Data quality / traceability", expanded=False):
        warnings = metadata.get("Warnings", [])
        if warnings:
            for warning in warnings:
                st.warning(warning)
        else:
            st.success("All expected core sections were detected.")

        service_cases = metadata.get("Service Cases", "")
        st.metric("Detected service cases", service_cases or "—")


# =========================================================
# TAB 2 — LOAD PARAMETER / TOWER LOADING
# =========================================================
with load_tab:
    load_parameters = dfs.get("Load Parameters", pd.DataFrame())
    antenna_height_resume = dfs.get("Antenna Height Resume", pd.DataFrame())
    antenna_loading = dfs.get("Antenna Loading", pd.DataFrame())
    linear_loading = dfs.get("Linear Loading", pd.DataFrame())

    st.markdown("### Load & design parameters")

    # Highlight the parameters that are most frequently used for design review.
    param_lookup: Dict[str, Any] = {}
    if not load_parameters.empty:
        for _, row in load_parameters.iterrows():
            param_lookup[str(row.get("Parameter", "")).upper()] = row.get("Value")

    p1, p2, p3, p4, p5 = st.columns(5)
    p1.metric("Design Code", param_lookup.get("CODE", "—"))
    vb = param_lookup.get("VB")
    p2.metric("Basic Wind (VB)", f"{vb} m/s" if vb not in (None, "") else "—")
    p3.metric("Structure Class", param_lookup.get("CLASS-G", "—"))
    p4.metric("Topographic Cat.", param_lookup.get("TOPCAT-G", "—"))
    p5.metric("North Angle", f"{param_lookup.get('ANGN')}°" if param_lookup.get("ANGN") not in (None, "") else "—")

    if not load_parameters.empty:
        st.dataframe(
            load_parameters,
            use_container_width=True,
            hide_index=True,
            height=min(430, 38 + (len(load_parameters) * 35)),
            column_config={
                "Parameter": st.column_config.TextColumn("Parameter", width="small"),
                "Value": st.column_config.TextColumn("Value", width="small"),
                "Description": st.column_config.TextColumn("Description", width="large"),
            },
        )
        st.caption(
            f"Load Parameters: {len(load_parameters):,} parameter(s). "
            "Exported to Excel sheet: Load Parameters."
        )
    else:
        st.info("PARAMETERS block was not detected in this file.")

    st.divider()
    st.markdown("### Antenna / device loading")

    if not antenna_loading.empty:
        a1, a2, a3, a4 = st.columns(4)
        a1.metric("Total Device", len(antenna_loading))
        a2.metric(
            "Existing",
            int((antenna_loading["Status"].astype(str) == "Existing").sum()),
        )
        a3.metric(
            "Proposed",
            int((antenna_loading["Status"].astype(str) == "Proposed").sum()),
        )
        mass_series = pd.to_numeric(antenna_loading["Added Mass"], errors="coerce")
        a4.metric(
            "Total Added Mass",
            f"{mass_series.sum():,.0f}" if mass_series.notna().any() else "—",
        )

        if not antenna_height_resume.empty:
            st.markdown("#### Resume device per elevation")
            resume_config = {
                "Elevation (m)": st.column_config.NumberColumn("Elevation", format="%d m"),
                "Total Device": st.column_config.NumberColumn("Total", format="%d"),
                "Device Summary": st.column_config.TextColumn("Device Summary", width="large"),
                "Tenant Device Summary": st.column_config.TextColumn("Per Tenant", width="large"),
                "Azimuth": st.column_config.TextColumn("Azimuth", width="medium"),
                "Status Summary": st.column_config.TextColumn("Status", width="medium"),
            }
            for device_col in antenna_height_resume.columns:
                if device_col not in {
                    "Elevation (m)", "Total Device", "Device Summary",
                    "Tenant Device Summary", "Azimuth", "Status Summary",
                }:
                    resume_config[device_col] = st.column_config.NumberColumn(
                        device_col, format="%d", width="small"
                    )

            st.dataframe(
                antenna_height_resume,
                use_container_width=True,
                hide_index=True,
                height=min(430, 38 + (min(len(antenna_height_resume), 10) * 35)),
                column_config=resume_config,
            )
            st.caption(
                "Elevation resume menggunakan band 1 meter berdasarkan ZA. "
                "Kolom numeric dibuat compact: RF, RRU, RRH, AAU, MW, ODU, dan Others. "
                "FILTER/TMA/COMBINER/DIPLEXER/TRIPLEXER/GPS tetap ditampilkan dengan nama asli "
                "di Device Summary dan Per Tenant. Azimuth menampilkan sudut unik per elevasi. "
                "Exported to Excel sheet: Antenna Height Resume."
            )

        st.markdown("#### Device detail")
        st.dataframe(
            antenna_loading,
            use_container_width=True,
            hide_index=True,
            height=min(620, 38 + (min(len(antenna_loading), 16) * 35)),
            column_config={
                "Status": st.column_config.TextColumn("Status", width="small"),
                "Load Group": st.column_config.TextColumn("Group", width="medium"),
                "Item": st.column_config.TextColumn("Item", width="small"),
                "Equipment Type": st.column_config.TextColumn("Type", width="small"),
                "XA (m)": st.column_config.NumberColumn("XA", format="%.3f"),
                "YA (m)": st.column_config.NumberColumn("YA", format="%.3f"),
                "ZA (m)": st.column_config.NumberColumn("ZA", format="%.2f"),
                "Angle (deg)": st.column_config.NumberColumn("Angle", format="%.0f°"),
                "Added Mass": st.column_config.NumberColumn("Mass", format="%.0f"),
            },
        )
        st.caption(
            f"Antenna Loading: {len(antenna_loading):,} device(s). "
            "Exported to Excel sheet: Antenna Loading."
        )
    else:
        st.info("ANCILLARIES antenna/device loading was not detected in this file.")

    st.divider()
    st.markdown("### Linear / additional loading")

    if not linear_loading.empty:
        type_counts = linear_loading["Load Type"].astype(str).value_counts()
        l1, l2, l3, l4 = st.columns(4)
        l1.metric("Total Linear Load", len(linear_loading))
        l2.metric("Ladder", int(type_counts.get("Ladder", 0)))
        l3.metric("Tray", int(type_counts.get("Tray", 0)))
        l4.metric("Feeder", int(type_counts.get("Feeder", 0)))

        st.dataframe(
            linear_loading,
            use_container_width=True,
            hide_index=True,
            height=min(500, 38 + (min(len(linear_loading), 12) * 35)),
            column_config={
                "Status": st.column_config.TextColumn("Status", width="small"),
                "Load Group": st.column_config.TextColumn("Group", width="medium"),
                "Item": st.column_config.TextColumn("Item", width="small"),
                "Load Type": st.column_config.TextColumn("Type", width="small"),
                "XB (m)": st.column_config.NumberColumn("XB", format="%.3f"),
                "YB (m)": st.column_config.NumberColumn("YB", format="%.3f"),
                "ZB (m)": st.column_config.NumberColumn("ZB", format="%.2f"),
                "XT (m)": st.column_config.NumberColumn("XT", format="%.3f"),
                "YT (m)": st.column_config.NumberColumn("YT", format="%.3f"),
                "ZT (m)": st.column_config.NumberColumn("ZT", format="%.2f"),
                "Factor": st.column_config.NumberColumn("Factor", format="%.2f"),
                "Angle (deg)": st.column_config.NumberColumn("Angle", format="%.0f°"),
            },
        )
        st.caption(
            f"Linear Loading: {len(linear_loading):,} item(s). "
            "Exported to Excel sheet: Linear Loading."
        )
    else:
        st.info("LINEAR LIBR loading (ladder, tray, feeder) was not detected in this file.")


# =========================================================
# TAB 3 — KETERANGAN MEMBER TOWER
# =========================================================
with member_tab:
    tower_profile = dfs.get("Tower Profile", pd.DataFrame())
    member_resume = dfs.get("Tower Member Resume", pd.DataFrame())
    member_detail = dfs.get("Tower Member Detail", pd.DataFrame())
    section_catalog = dfs.get("Member Section Catalog", pd.DataFrame())
    bolt_data = dfs.get("Bolt Data", pd.DataFrame())

    st.markdown("### Tower profile & member mapping")

    profile_row: Dict[str, Any] = {}
    if not tower_profile.empty:
        profile_row = tower_profile.iloc[0].to_dict()

    title_1 = str(profile_row.get("Title 1", "") or "")
    title_2 = str(profile_row.get("Title 2", "") or "")
    if title_1 or title_2:
        st.caption(" | ".join(x for x in [title_1, title_2] if x))

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Panels", int(profile_row.get("Total Panels", 0) or 0))
    faces = profile_row.get("Faces")
    m2.metric("Faces", f"{faces:g}" if isinstance(faces, (int, float)) and pd.notna(faces) else "—")
    wbase = profile_row.get("WBASE")
    m3.metric("Base Width", f"{wbase:g} m" if isinstance(wbase, (int, float)) and pd.notna(wbase) else "—")
    m4.metric("Section Types", int(profile_row.get("Section Types", 0) or 0))
    m5.metric("Bolt Types", len(bolt_data))

    # Build geometry once. The three UI views below share this exact model so
    # 2D, 3D and validation cannot silently diverge.
    geometry_model = None
    geometry_error = None
    if selected_content:
        try:
            geometry_model = build_tower_geometry(
                selected_content,
                section_catalog=section_catalog,
                udp_sources=udp_sources,
            )
        except Exception as geometry_exc:
            geometry_error = str(geometry_exc)

    visual_tab, member_data_tab = st.tabs(
        ["🧭 Visualisasi Geometri", "📋 Member & Material Data"]
    )

    # -----------------------------------------------------
    # GEOMETRY VISUALIZATION — 2D + material + 3D visible together.
    # -----------------------------------------------------
    with visual_tab:
        if geometry_error:
            st.error(f"Geometry tower belum dapat dibentuk: {geometry_error}")
        elif geometry_model is None:
            st.info("Source MST/RPT/TXT belum tersedia untuk membentuk geometry tower.")
        else:
            validation = geometry_validation_summary(geometry_model)
            validation_df = geometry_validation_dataframe(geometry_model)

            if int(geometry_model.faces) >= 3:
                side_options = [f"Face {i}" for i in range(1, int(geometry_model.faces) + 1)]
            else:
                side_options = ["Side X", "Side Y"]

            # Compact health bar: important information remains visible without
            # pushing the actual geometry too far down the page.
            status_text = {
                "PASS": "✅ Complete",
                "PARTIAL": "⚠️ Partial",
                "REVIEW": "⚠️ Review",
                "NO DATA": "— No data",
            }.get(validation.get("status"), str(validation.get("status", "—")))

            h1, h2, h3, h4, h5 = st.columns([1.0, 1.0, 1.0, 1.0, 1.15])
            h1.metric("Geometry", status_text)
            h2.metric("Completeness", f"{validation.get('completeness', 0):.0f}%")
            h3.metric("Incomplete", validation.get("incomplete", 0))
            h4.metric("Unsupported", validation.get("unsupported", 0))
            udp_calls = validation.get("udp_calls", 0)
            udp_resolved = validation.get("udp_resolved", 0)
            h5.metric("UDP @reference", f"{udp_resolved}/{udp_calls}" if udp_calls else "—")

            if validation.get("zero_violations", 0):
                st.error(
                    f"Terdeteksi {validation['zero_violations']} zero-member violation. "
                    "Member dengan section 0 tidak boleh masuk geometry."
                )

            # Two principal visual areas: the engineering elevation on the left
            # and the interactive 3D model on the right. The material schedule
            # sits directly beside the 2D tower inside the left area, matching
            # conventional tower outline drawings without hiding the 3D view.
            left_visual, right_visual = st.columns([1.08, 1.0], gap="large", vertical_alignment="top")

            with left_visual:
                st.markdown("#### 2D Engineering View")
                geometry_view = st.selectbox(
                    "Sisi tampak",
                    side_options,
                    index=0,
                    key=f"geometry_view_{selected_name}",
                    help=(
                        "2D dibuat statis seperti outline drawing. Hanya sisi yang dipilih berubah; "
                        "member dengan section <= 0 selalu dihilangkan."
                    ),
                )

                tower_col, schedule_col = st.columns([0.92, 1.08], gap="small")
                with tower_col:
                    fig_2d = make_engineering_2d_figure(
                        geometry_model,
                        view=geometry_view,
                        height_px=760,
                    )
                    st.plotly_chart(
                        fig_2d,
                        use_container_width=True,
                        key=f"tower_engineering_2d_{selected_name}_{geometry_view}",
                        config={
                            "staticPlot": True,
                            "displayModeBar": False,
                            "displaylogo": False,
                            "scrollZoom": False,
                        },
                    )

                with schedule_col:
                    fig_material = make_material_elevation_figure(
                        geometry_model,
                        height_px=760,
                    )
                    st.plotly_chart(
                        fig_material,
                        use_container_width=True,
                        key=f"tower_material_schedule_{selected_name}_{geometry_view}",
                        config={
                            "staticPlot": True,
                            "displayModeBar": False,
                            "displaylogo": False,
                            "scrollZoom": False,
                        },
                    )

                st.caption(
                    "2D menggunakan elevasi/Z asli. Lebar hanya diperbesar secara konstan untuk keterbacaan "
                    "outline; taper antar-section tetap konsisten. Material identik yang berurutan digabung "
                    "agar label tidak bertumpuk."
                )

            with right_visual:
                st.markdown("#### 3D Interactive View")
                available_families = [
                    family for family in GEOMETRY_FAMILY_ORDER
                    if any(m.family == family for m in geometry_model.members)
                ]
                extra_geometry_families = sorted(
                    {m.family for m in geometry_model.members} - set(available_families)
                )
                available_families += extra_geometry_families

                geometry_families = st.multiselect(
                    "Member pada gambar 3D",
                    options=available_families,
                    default=available_families,
                    key=f"geometry_family_{selected_name}",
                    help="LEG, BR, HOR, RED, PBR, HIP, CROSS dan UDP dapat diaktifkan/nonaktifkan.",
                )
                geometry_panel_options = sorted({int(p.panel_no) for p in geometry_model.panels})
                geometry_panels = st.multiselect(
                    "Panel filter 3D",
                    options=geometry_panel_options,
                    default=[],
                    placeholder="All panels",
                    key=f"geometry_panel_{selected_name}",
                )

                fig_3d = make_tower_3d_figure(
                    geometry_model,
                    families=geometry_families,
                    panel_numbers=geometry_panels or None,
                    height_px=760,
                )
                st.plotly_chart(
                    fig_3d,
                    use_container_width=True,
                    key=f"tower_3d_{selected_name}",
                    config={"displaylogo": False, "scrollZoom": True},
                )
                st.caption(
                    f"Height {geometry_model.total_height:.2f} m | Base width {geometry_model.wbase:.2f} m | "
                    f"{geometry_model.faces}-leg | {len(geometry_model.panels)} panel(s) | "
                    f"{len(geometry_model.members):,} visual member segment(s). "
                    "Aspect 3D disesuaikan untuk keterbacaan; koordinat engineering tidak diubah."
                )

            udp_member_count = sum(1 for m in geometry_model.members if str(m.family).upper() == "UDP")
            if udp_sources:
                if udp_member_count:
                    st.caption(
                        f"UDP coordinate geometry: {len(udp_sources):,} source(s) tersedia | "
                        f"{udp_member_count:,} NODE/MEMB segment(s) masuk ke model visual."
                    )
                else:
                    st.warning(
                        "UDP source tersedia tetapi belum ada NODE/MEMB segment yang masuk ke geometry. "
                        "Periksa @reference dan internal UDP name pada Geometry diagnostics."
                    )

            material_schedule = material_schedule_dataframe(geometry_model)
            detail_col, diag_col = st.columns([1.0, 1.0])
            with detail_col:
                with st.expander("Material schedule detail", expanded=False):
                    if not material_schedule.empty:
                        st.dataframe(
                            material_schedule,
                            use_container_width=True,
                            hide_index=True,
                            height=min(540, 40 + min(len(material_schedule), 13) * 35),
                        )
                    else:
                        st.info("Material schedule belum dapat dibentuk dari geometry yang terdeteksi.")

            with diag_col:
                with st.expander("Geometry diagnostics", expanded=False):
                    if not validation_df.empty:
                        st.markdown("**Assignment vs plotted geometry**")
                        st.dataframe(
                            validation_df,
                            use_container_width=True,
                            hide_index=True,
                            height=min(480, 40 + min(len(validation_df), 11) * 36),
                            column_config={
                                "Panel": st.column_config.NumberColumn("Panel", format="%d", width="small"),
                                "Completeness (%)": st.column_config.NumberColumn("Complete", format="%.0f%%", width="small"),
                                "Status": st.column_config.TextColumn("Status", width="small"),
                            },
                        )

                    if geometry_model.warnings:
                        st.markdown("**Geometry notes**")
                        for geometry_warning in geometry_model.warnings:
                            st.warning(geometry_warning)

                    panel_geometry_df = geometry_panels_dataframe(geometry_model)
                    if not panel_geometry_df.empty:
                        st.markdown("**Interpolated panel geometry**")
                        st.dataframe(
                            panel_geometry_df,
                            use_container_width=True,
                            hide_index=True,
                            height=min(360, 38 + min(len(panel_geometry_df), 8) * 35),
                        )

                    member_geometry_df = geometry_members_dataframe(geometry_model)
                    if not member_geometry_df.empty:
                        st.markdown("**Generated visual member segments**")
                        st.dataframe(
                            member_geometry_df,
                            use_container_width=True,
                            hide_index=True,
                            height=320,
                        )

    # -----------------------------------------------------
    # MEMBER / MATERIAL DATA — stable tables moved into one dedicated sub-tab.
    # -----------------------------------------------------
    with member_data_tab:
        if not member_resume.empty:
            st.markdown("#### Resume member per panel")
            st.dataframe(
                member_resume,
                use_container_width=True,
                hide_index=True,
                height=min(680, 38 + (min(len(member_resume), 18) * 35)),
                column_config={
                    "Panel": st.column_config.NumberColumn("Panel", format="%d", width="small"),
                    "Elevation Range": st.column_config.TextColumn("Elevation", width="small"),
                    "HT (m)": st.column_config.NumberColumn("HT", format="%.2f", width="small"),
                    "TW": st.column_config.NumberColumn("TW", format="%.2f", width="small"),
                    "FACE Pattern": st.column_config.TextColumn("FACE", width="small"),
                    "PLAN Pattern": st.column_config.TextColumn("PLAN", width="small"),
                    "HIP Pattern": st.column_config.TextColumn("HIP Pattern", width="small"),
                    "LEG": st.column_config.TextColumn("LEG", width="large"),
                    "BR": st.column_config.TextColumn("BR", width="large"),
                    "HOR": st.column_config.TextColumn("HOR", width="large"),
                    "RED": st.column_config.TextColumn("RED", width="large"),
                    "PBR": st.column_config.TextColumn("PBR", width="large"),
                    "HIP": st.column_config.TextColumn("HIP", width="large"),
                },
            )
            st.caption(
                "Setiap kode PROFILE dipetakan ke master SECTIONS. Member dengan section ID 0 "
                "dianggap tidak terpasang dan tidak dimasukkan ke detail maupun geometry."
            )
        else:
            st.info("PROFILE member assignment belum terdeteksi pada file ini.")

        if not member_detail.empty:
            st.markdown("#### Detail member tower")
            family_options = ["All"] + [
                x for x in ["LEG", "BR", "HOR", "RED", "PBR", "HIP"]
                if x in set(member_detail["Member Family"].astype(str))
            ]
            extra_families = sorted(set(member_detail["Member Family"].astype(str)) - set(family_options))
            family_options += extra_families

            fcol, pcol = st.columns([1.0, 1.0])
            with fcol:
                selected_family = st.selectbox(
                    "Member family", family_options, key=f"member_family_{selected_name}"
                )
            with pcol:
                panel_options = ["All"] + [str(int(x)) for x in sorted(member_detail["Panel"].dropna().unique())]
                selected_panel = st.selectbox(
                    "Panel", panel_options, key=f"member_panel_{selected_name}"
                )

            member_view = member_detail.copy()
            if selected_family != "All":
                member_view = member_view[member_view["Member Family"].astype(str) == selected_family]
            if selected_panel != "All":
                member_view = member_view[member_view["Panel"] == int(selected_panel)]

            st.dataframe(
                member_view,
                use_container_width=True,
                hide_index=True,
                height=min(620, 38 + (min(len(member_view), 16) * 35)),
                column_config={
                    "Panel": st.column_config.NumberColumn("Panel", format="%d", width="small"),
                    "Section From (m)": None,
                    "Section To (m)": None,
                    "Elevation Range": st.column_config.TextColumn("Elevation", width="small"),
                    "Panel HT (m)": st.column_config.NumberColumn("HT", format="%.2f", width="small"),
                    "TW": st.column_config.NumberColumn("TW", format="%.2f", width="small"),
                    "Line Type": st.column_config.TextColumn("Source", width="small"),
                    "Pattern": st.column_config.TextColumn("Pattern", width="small"),
                    "Qualifier": st.column_config.TextColumn("Position", width="small"),
                    "Member Family": st.column_config.TextColumn("Family", width="small"),
                    "Member Key": st.column_config.TextColumn("Member", width="small"),
                    "Section ID": st.column_config.NumberColumn("Section ID", format="%d", width="small"),
                    "Section / Profile": st.column_config.TextColumn("Section / Profile", width="medium"),
                    "fy": st.column_config.NumberColumn("fy", format="%.0f", width="small"),
                    "Section Category": st.column_config.TextColumn("Category", width="small"),
                    "Mapped": st.column_config.TextColumn("Mapped", width="small"),
                },
            )
            st.caption(
                f"Tower Member Detail: {len(member_detail):,} member assignment(s). "
                "Exported to Excel sheet: Tower Member Detail."
            )

        st.divider()
        catalog_col, bolt_col = st.columns([1.15, 1.0])
        with catalog_col:
            st.markdown("#### Master section member")
            if not section_catalog.empty:
                st.dataframe(
                    section_catalog,
                    use_container_width=True,
                    hide_index=True,
                    height=min(520, 38 + (min(len(section_catalog), 13) * 35)),
                    column_config={
                        "Section ID": st.column_config.NumberColumn("ID", format="%d", width="small"),
                        "Section / Profile": st.column_config.TextColumn("Section / Profile", width="medium"),
                        "fy": st.column_config.NumberColumn("fy", format="%.0f", width="small"),
                        "Category": st.column_config.TextColumn("Category", width="small"),
                        "Library": st.column_config.TextColumn("Library", width="small"),
                        "IFACT": st.column_config.NumberColumn("IFACT", format="%.2f", width="small"),
                    },
                )
                st.caption("Exported to Excel sheet: Member Section Catalog.")
            else:
                st.info("SECTIONS member catalog tidak terdeteksi.")

        with bolt_col:
            st.markdown("#### Bolt data")
            if not bolt_data.empty:
                st.dataframe(
                    bolt_data,
                    use_container_width=True,
                    hide_index=True,
                    height=min(520, 38 + (min(len(bolt_data), 13) * 35)),
                    column_config={
                        "Bolt ID": st.column_config.TextColumn("Bolt ID", width="small"),
                        "Grade": st.column_config.TextColumn("Grade", width="small"),
                        "Diameter (mm)": st.column_config.NumberColumn("D (mm)", format="%.0f", width="small"),
                        "AS": st.column_config.NumberColumn("AS", format="%.0f", width="small"),
                        "FY": st.column_config.NumberColumn("FY", format="%.0f", width="small"),
                        "FU": st.column_config.NumberColumn("FU", format="%.0f", width="small"),
                        "FV": st.column_config.NumberColumn("FV", format="%.0f", width="small"),
                        "NSP": st.column_config.NumberColumn("NSP", format="%.0f", width="small"),
                        "Remark": st.column_config.TextColumn("Remark", width="large"),
                    },
                )
                st.caption(
                    f"Bolt Data: {len(bolt_data):,} bolt definition(s). Exported to Excel sheet: Bolt Data."
                )
            else:
                st.info("BOLTDATA tidak terdeteksi pada file ini.")

st.caption(
    "Engineering note: this application is an extraction and screening tool. "
    "PASS/FAIL shown here must remain traceable to the governing design criteria "
    "and the full MS Tower analysis report."
)
